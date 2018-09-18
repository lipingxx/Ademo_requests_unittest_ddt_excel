# coding:utf-8
from openpyxl import load_workbook
import openpyxl

def copy_excel(excelpath1, excelpath2):
    '''
    复制excel，把excelpath1数据复制到excelpath2
    '''
    # 写入excel之前，先创建worksheet对象
    wb2 = openpyxl.Workbook()
    # 保存文件
    wb2.save(excelpath2)
    # load_workbook(filename):读取数据
    wb1 = openpyxl.load_workbook(excelpath1)
    wb2 = openpyxl.load_workbook(excelpath2)
    # 获取工作簿中所有表格的名称，并以列表的形式返回，['sheet1','sheet2']
    # ！！！workbook对象提供了很多属性和方法！！！
    sheets1 = wb1.sheetnames
    sheets2 = wb2.sheetnames
    sheet1 = wb1[sheets1[0]]
    sheet2 = wb2[sheets2[0]]
    # 原表中的最大行数、最大列数
    max_row = sheet1.max_row
    max_column = sheet1.max_column
    # m行 n列
    for m in list(range(1,max_row+1)):  # 0 ~ m+1
        for n in list(range(97,97+max_column)):   # chr(97)='a',从A列开始 ~ max_column
            n = chr(n)                            # ASCII字符
            i ='%s%d'% (n, m)                     # 单元格编号
            cell1 = sheet1[i].value               # 获取data单元格数据  sheet(n,m)
            sheet2[i].value = cell1               # 赋值到test单元格

    # 保存数据
    wb2.save(excelpath2)
    # 关闭excel
    wb1.close()
    wb2.close()

class Write_excel(object):
    '''修改excel数据'''
    def __init__(self, filename):
        self.filename = filename
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active  # 激活sheet

    def write(self, row_n, col_n, value):
        '''写入数据，如(2,3，"hello"),第二行第三列写入数据"hello"'''
        self.ws.cell(row_n, col_n).value = value
        self.wb.save(self.filename)

'''
if __name__ == "__main__":
    copy_excel("debug_api.xlsx", "test111.xlsx")
    wt = Write_excel("test111.xlsx")
    wt.write(4, 5, "HELLEOP")
    wt.write(4, 6, "HELLEOP")
'''


